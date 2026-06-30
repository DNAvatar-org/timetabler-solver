# File: charger.py - Chargement Excel et dict SheetJS → DonneesCollege
# Desc: En français, dans l'architecture emploi du temps, je suis…
# Version 1.0.0
# Copyright 2025 DNAvatar.org - Arnaud Maignan
# Licensed under Apache License 2.0 with Commons Clause.
# See LICENSE_HEADER.txt for full terms.
# Date: June 27, 2025
# Logs:
# - Split from solve_ecoles.py monolith

from __future__ import annotations
from pathlib import Path
import openpyxl

from solver.constants import JOURS, XLSX
from solver.donnees import (
    Classe, Creneau, DistanceEtab, DonneesCollege, Etablissement,
    Matiere, Preference, Prof, ProgrammeItem, Salle,
)
from solver.programme import _split_option_header

def charger_donnees(path: Path = XLSX) -> DonneesCollege:
    wb = openpyxl.load_workbook(path, data_only=True)
    d = DonneesCollege()

    ws = wb["classes"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        id_, etab_id, nom, niveau, nb, spec = row
        d.classes.append(Classe(id_, nom, niveau, int(float(nb or 0)), spec, str(etab_id or "")))

    ws = wb["matieres"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if not vals[0]: continue
        id_, nom, abrev, salle_type, duree = vals[:5]
        couleur = str(vals[5]).strip() if len(vals) > 5 and vals[5] else ""
        d.matieres[id_] = Matiere(id_, nom, abrev, salle_type, int(duree or 1), couleur)

    ws = wb["programmes"]
    prog_header = [str(c.value or '') for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if prog_header and prog_header[0].strip().lower() == "matiere_id":
        known_niveaux = {c.niveau for c in d.classes}
        all_cols = prog_header[2:]
        for row in ws.iter_rows(min_row=2, values_only=True):
            mat_id = str(row[0]).strip() if row[0] not in (None, "") else ""
            if not mat_id: continue
            for j, col_name in enumerate(all_cols):
                ci = j + 2
                h_val = row[ci] if ci < len(row) else None
                if h_val:
                    try:
                        h = float(h_val)
                        if h > 0:
                            if col_name in known_niveaux:
                                d.programme.append(ProgrammeItem(col_name, mat_id, h, None))
                            else:
                                niv, opt = _split_option_header(col_name, known_niveaux)
                                d.programme.append(ProgrammeItem(niv, mat_id, h, None, opt))
                    except (ValueError, TypeError): pass
    else:
        for row in ws.iter_rows(min_row=2, values_only=True):
            _etab, niveau, matiere_id, _nom, h_semaine, _h_annee, _nb, notes = row
            if h_semaine:
                d.programme.append(ProgrammeItem(niveau, matiere_id, float(h_semaine), notes))

    ws = wb["professeurs"]
    hdr_p = [str(c.value or "").strip() for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = {hdr_p[i]: row[i] for i in range(min(len(hdr_p), len(row)))}
        id_ = v.get("id")
        if not id_: continue
        matieres = [m.strip() for m in (v.get("matieres") or "").split(",") if m.strip()]
        niveaux = {n.strip() for n in (v.get("niveaux") or "").split(",") if n.strip()}
        h_contrat = v.get("h_contrat")
        d.profs[id_] = Prof(id_, f"{v.get('prenom','')} {v.get('nom','')}", matieres, int(h_contrat or 18), set(JOURS), niveaux)

    ws = wb["salles"]
    hdr_s = [str(c.value or "").strip() for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        vals = {hdr_s[i]: row[i] for i in range(min(len(hdr_s), len(row)))}
        id_ = str(vals.get("id", row[0])).strip()
        etab_id = str(vals.get("etab_id", "") or "")
        nom = vals.get("nom", "")
        type_ = vals.get("type", "standard")
        capacite = vals.get("capacite", 0)
        stabilite = (
            vals.get("stabilité")
            or vals.get("stabilite")
            or "partagee"
        )
        stab = str(stabilite or "partagee").strip() or "partagee"
        d.salles.append(Salle(
            id_, nom, str(type_ or "standard"), int(capacite or 0), etab_id, stab,
        ))

    if "etablissements" in wb.sheetnames:
        ws = wb["etablissements"]
        hdr = [str(c.value or "").strip() for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            vals = {hdr[i]: row[i] for i in range(min(len(hdr), len(row)))}
            eid = str(vals.get("id", row[0])).strip()
            d.etablissements[eid] = Etablissement(eid)

    ref_name = "referentiel" if "referentiel" in wb.sheetnames else "domaines"
    ws = wb[ref_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        onglet, col, valeurs = (row + (None,))[:3]
        if onglet == "horaires" and col and valeurs is not None:
            d.ref[str(col)] = str(valeurs).strip()

    ws = wb["creneaux"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        id_, jour, debut, fin, type_ = row
        type_ = str(type_ or "cours")
        cr = Creneau(id_, jour, debut, fin, type_)
        d.horaires.append(cr)
        if type_ == "cours":
            d.creneaux.append(cr)

    ws = wb["preferences"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if not vals[0]:
            continue
        prof_id, type_, valeur, priorite = vals[:4]
        operateur = str(vals[5] or '').strip() if len(vals) > 5 else ''
        d.preferences.append(Preference(
            prof_id, type_, str(valeur or ''), str(priorite or ''), operateur or 'ET'
        ))

    if "distances" in wb.sheetnames:
        ws = wb["distances"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = list(row) + [None] * 5
            trajet, dist_km, min_transport, reg_demi, reg_jour = vals[:5]
            if not trajet:
                continue
            parts = str(trajet).split("-")
            if len(parts) == 2:
                d.distances_etab.append(DistanceEtab(
                    parts[0].strip(), parts[1].strip(),
                    float(dist_km or 0), int(float(min_transport or 0)),
                    str(reg_demi or "").strip(), str(reg_jour or "").strip(),
                ))

    # jour_libre contrainte → indisponibilité dure (remplace les colonnes lundi-vendredi)
    for pref in d.preferences:
        if pref.type == "jour_libre" and pref.priorite == "contrainte":
            if pref.prof_id in d.profs:
                d.profs[pref.prof_id].jours_dispo.discard(pref.valeur)

    return d


def charger_donnees_dict(sheets: dict) -> DonneesCollege:
    """
    Construit DonneesCollege depuis un dict {nom_feuille: {columns, rows}}.
    Format identique à xls_to_json() — toutes les valeurs sont des strings.
    Compatible avec SheetJS côté client.
    """
    d = DonneesCollege()

    def rows_of(nom: str) -> list:
        s = sheets.get(nom, {})
        return [r for r in s.get("rows", []) if any(v for v in r)]

    def s2i(v, default=0) -> int:
        try: return int(float(v)) if v not in (None, "") else default
        except (ValueError, TypeError): return default

    def s2f(v, default=0.0) -> float:
        try: return float(v) if v not in (None, "") else default
        except (ValueError, TypeError): return default

    def s2s(v) -> str | None:
        s = str(v).strip() if v not in (None, "") else ""
        return s if s else None

    for row in rows_of("classes"):
        row = list(row) + [""] * 6
        id_, etab_id, nom, niveau, nb, spec = row[:6]
        if not id_: continue
        d.classes.append(Classe(id_, nom, niveau, s2i(nb), s2s(spec), str(etab_id or "")))

    for row in rows_of("matieres"):
        row = list(row) + [""] * 6
        id_, nom, abrev, salle_type, duree, couleur = row[:6]
        if not id_: continue
        d.matieres[id_] = Matiere(id_, nom, abrev, s2s(salle_type), s2i(duree, 1), str(couleur).strip())

    prog_sheet = sheets.get("programmes", {})
    prog_cols  = prog_sheet.get("columns", [])
    if prog_cols and str(prog_cols[0]).strip().lower() == "matiere_id":
        # Format matrice : [matiere_id, nom, niveau1, niveau2, ..., option1, option2, ...]
        known_niveaux = {c.niveau for c in d.classes}
        all_cols = [str(c) for c in prog_cols[2:] if c not in (None, "")]
        for row in rows_of("programmes"):
            mat_id = str(row[0]).strip() if len(row) > 0 and row[0] not in (None, "") else ""
            if not mat_id:
                continue
            for j, col_name in enumerate(all_cols):
                ci = j + 2
                h = s2f(row[ci] if ci < len(row) else "")
                if h > 0:
                    if col_name in known_niveaux:
                        d.programme.append(ProgrammeItem(col_name, mat_id, h, None))
                    else:
                        niv, opt = _split_option_header(col_name, known_niveaux)
                        d.programme.append(ProgrammeItem(niv, mat_id, h, None, opt))
    else:
        # Format liste plate (legacy)
        for row in rows_of("programmes"):
            row = list(row) + [""] * 8
            _etab, niveau, matiere_id, _nom, h_semaine = row[:5]
            notes = row[7] if len(row) > 7 else None
            if matiere_id and s2f(h_semaine):
                d.programme.append(ProgrammeItem(niveau, matiere_id, s2f(h_semaine), s2s(notes)))

    _prof_cols = sheets.get("professeurs", {}).get("columns", [])

    def _cval(row: list, name: str, default: str = "") -> str:
        try:
            i = _prof_cols.index(name)
            return str(row[i]) if i < len(row) and row[i] not in (None, "") else default
        except ValueError:
            return default

    for row in rows_of("professeurs"):
        row = list(row) + [""] * 16
        id_ = _cval(row, "id") or row[0]
        if not id_: continue
        nom          = _cval(row, "nom")
        prenom       = _cval(row, "prenom")
        matieres_str = _cval(row, "matieres")
        niveaux_str  = _cval(row, "niveaux")
        h_contrat    = _cval(row, "h_contrat", "18")
        matieres = [m.strip() for m in str(matieres_str or "").split(",") if m.strip()]
        niveaux = {n.strip() for n in str(niveaux_str or "").split(",") if n.strip()}
        d.profs[id_] = Prof(id_, f"{prenom} {nom}", matieres, s2i(h_contrat, 18), set(JOURS), niveaux)

    _salle_cols = sheets.get("salles", {}).get("columns", [])

    def _salleval(row: list, name: str, default: str = "") -> str:
        try:
            i = _salle_cols.index(name)
            return str(row[i]).strip() if i < len(row) and row[i] not in (None, "") else default
        except ValueError:
            return default

    for row in rows_of("salles"):
        row = list(row) + [""] * 8
        id_ = _salleval(row, "id") or row[0]
        if not id_:
            continue
        stabilite = (
            _salleval(row, "stabilité")
            or _salleval(row, "stabilite", "partagee")
            or "partagee"
        )
        d.salles.append(Salle(
            id_,
            _salleval(row, "nom"),
            _salleval(row, "type", "standard"),
            s2i(_salleval(row, "capacite", "0")),
            _salleval(row, "etab_id"),
            stabilite,
        ))

    _etab_cols = sheets.get("etablissements", {}).get("columns", [])

    def _etabval(row: list, name: str, default: str = "") -> str:
        try:
            i = _etab_cols.index(name)
            return str(row[i]).strip() if i < len(row) and row[i] not in (None, "") else default
        except ValueError:
            return default

    for row in rows_of("etablissements"):
        row = list(row) + [""] * 20
        eid = _etabval(row, "id") or str(row[0]).strip()
        if not eid:
            continue
        d.etablissements[eid] = Etablissement(eid)

    for row in rows_of("referentiel") or rows_of("domaines"):
        row = list(row) + ["", "", ""]
        onglet, col, valeurs = row[0], row[1], row[2]
        if onglet == "horaires" and col and valeurs:
            d.ref[str(col)] = str(valeurs).strip()

    for row in rows_of("creneaux"):
        row = list(row) + [""] * 5
        id_, jour, debut, fin, type_ = row[:5]
        if not id_: continue
        type_ = str(type_ or "cours")
        cr = Creneau(id_, jour, str(debut), str(fin), type_)
        d.horaires.append(cr)
        if type_ == "cours":
            d.creneaux.append(cr)

    for row in rows_of("preferences"):
        row = list(row) + [""] * 6
        prof_id, type_, valeur, priorite = row[:4]
        if not prof_id: continue
        operateur = str(row[5] or "").strip()
        d.preferences.append(Preference(
            prof_id, type_, str(valeur or ""), str(priorite or ""), operateur or "ET"
        ))

    for row in rows_of("distances"):
        row = list(row) + [""] * 5
        trajet, dist_km, min_transport, reg_demi, reg_jour = row[:5]
        if not trajet:
            continue
        parts = str(trajet).split("-")
        if len(parts) == 2:
            d.distances_etab.append(DistanceEtab(
                parts[0].strip(), parts[1].strip(),
                float(dist_km) if dist_km else 0.0,
                int(float(min_transport)) if min_transport else 0,
                str(reg_demi).strip(), str(reg_jour).strip(),
            ))

    # jour_libre contrainte → indisponibilité dure (remplace les colonnes lundi-vendredi)
    for pref in d.preferences:
        if pref.type == "jour_libre" and pref.priorite == "contrainte":
            if pref.prof_id in d.profs:
                d.profs[pref.prof_id].jours_dispo.discard(pref.valeur)

    return d
