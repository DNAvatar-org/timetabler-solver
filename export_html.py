"""
Génère un fichier HTML autonome affichant l'emploi du temps optimisé.
Grille par classe (onglets) + vue par professeur + rapport de contraintes.
"""

from __future__ import annotations
from pathlib import Path
from collections import defaultdict
import base64 as _base64
import csv as _csv
import html as _html
import io
import urllib.parse

COULEURS: dict[str, tuple[str, str]] = {
    "MATH": ("#4e79a7", "#fff"),
    "FR":   ("#59a14f", "#fff"),
    "HG":   ("#f28e2b", "#fff"),
    "ANG":  ("#76b7b2", "#fff"),
    "ESP":  ("#b07aa1", "#fff"),
    "ALL":  ("#9467bd", "#fff"),
    "SVT":  ("#17becf", "#fff"),
    "PC":   ("#1f77b4", "#fff"),
    "TECH": ("#8c564b", "#fff"),
    "EPS":  ("#e15759", "#fff"),
    "AP":   ("#edc948", "#333"),
    "MUS":  ("#ff9da7", "#333"),
    "EMC":  ("#d4a574", "#333"),
    "LAT":  ("#bcbd22", "#333"),
}
DEFAULT_COULEUR = ("#aaa", "#fff")
LIBELLE_COMBLEUR = "Combleur fictif"
TOOLTIP_COMBLEUR_ALT0 = "Prof fictif — comble le déficit d'heures"
TOOLTIP_COMBLEUR_ALT2 = (
    "Placeholder de staffing : les vrais profs ne suffisent pas. "
    "Réduire les Combleurs ou recruter."
)
JOURS = ["lundi", "mardi", "mercredi", "jeudi", "vendredi"]
JOURS_LABELS = {
    "lundi": "Lundi", "mardi": "Mardi", "mercredi": "Mercredi",
    "jeudi": "Jeudi", "vendredi": "Vendredi",
}


def _est_combleur_nom(nom: str) -> bool:
    return "combleur" in (nom or "").lower()


def _html_nom_prof(prof) -> str:
    """Libellé affiché : Combleur fictif (rouge + alt0sec) ou nom réel."""
    if prof is None:
        return ""
    nom = prof.nom_complet if hasattr(prof, "nom_complet") else str(prof)
    if not _est_combleur_nom(nom):
        return nom
    tip0 = TOOLTIP_COMBLEUR_ALT0.replace('"', "&quot;")
    tip2 = TOOLTIP_COMBLEUR_ALT2.replace('"', "&quot;")
    return (
        f'<span class="prof-combleur" data-tooltip="{tip0}" data-alt2sec="{tip2}">'
        f"{LIBELLE_COMBLEUR}</span>"
    )


def _unpack_cours(entry) -> tuple[str, str, str | None]:
    m_id, p_id = entry[0], entry[1]
    s_id = entry[2] if len(entry) >= 3 and entry[2] else None
    return m_id, p_id, s_id


def _nom_salle(d, c_id: str, m_id: str, s_id: str | None) -> str:
    if s_id:
        for s in d.salles:
            if s.id == s_id:
                return s.nom
        return s_id
    classe = next((c for c in d.classes if c.id == c_id), None)
    if not classe:
        return ""
    mat = d.matieres.get(m_id)
    if not mat:
        return ""
    st = mat.salle_type or "standard"
    for s in d.salles:
        if s.etab_id == classe.etab_id and s.type == st:
            return s.nom
    return ""


def _libelle_salle_affichage(nom: str) -> str:
    """Affichage : « Salle 101 » → « 101 », sans le mot Salle."""
    if not nom:
        return ""
    n = nom.strip()
    if n.lower().startswith("salle "):
        return n[6:].strip()
    return n


# Export XLSX : une seule police par cellule (pas de rich text OpenOffice)
_XLSX_FONT_SIZE_CLASSE = 8   # matière + prof + salle sur 3 lignes
_XLSX_FONT_SIZE_PROF = 8
_XLSX_ROW_HEIGHT_CLASSE = 36
_XLSX_ROW_HEIGHT_PROF = 32


def _xlsx_font(color: str, *, bold: bool = False, size: int = _XLSX_FONT_SIZE_CLASSE) -> "Font":
    from openpyxl.styles import Font
    return Font(bold=bold, color=color, size=size)

MATIN_DEBUTS = ("08:00", "09:00", "10:15", "11:15")
APREM_DEBUTS = ("13:15", "14:15", "15:30", "16:30")


def _auto_fg(hex_bg: str) -> str:
    h = hex_bg.lstrip("#")
    if len(h) != 6:
        return "#fff"
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return "#fff" if (0.299 * r + 0.587 * g + 0.114 * b) / 255 < 0.6 else "#333"


def _excel_rgb(hex_color: str) -> str:
    """openpyxl : hex 6 ou 8 caractères sans # (#fff → FFFFFF)."""
    h = str(hex_color).strip().lstrip("#").upper()
    if len(h) == 3:
        h = "".join(c * 2 for c in h)
    if len(h) == 8:
        return h
    if len(h) == 6:
        try:
            int(h, 16)
            return h
        except ValueError:
            pass
    return "FFFFFF"


def _build_couleurs(d) -> dict[str, tuple[str, str]]:
    c = dict(COULEURS)
    for m in d.matieres.values():
        if m.couleur:
            c[m.abrev] = (m.couleur, _auto_fg(m.couleur))
    return c


def _couleur(abrev: str, couleurs: dict) -> tuple[str, str]:
    return couleurs.get(abrev, DEFAULT_COULEUR)


def _fmt_horaire(h: str) -> str:
    return h.replace(":", "h", 1)


def _plages_dejeuner(d) -> list[tuple[str, str]]:
    raw = d.ref.get("creneaux_dejeuner", "")
    if raw:
        plages: list[tuple[str, str]] = []
        for part in raw.split("|"):
            part = part.strip()
            if "-" not in part:
                continue
            debut, fin = part.split("-", 1)
            plages.append((debut.strip(), fin.strip()))
        if plages:
            return plages[:2]
    # Dédupliquer par debut uniquement (fin peut varier selon le jour)
    vus_debut: set[str] = set()
    plages = []
    for t in sorted(d.horaires, key=lambda t: t.fin, reverse=True):
        if t.type != "dejeuner":
            continue
        if t.debut not in vus_debut:
            vus_debut.add(t.debut)
            plages.append((t.debut, t.fin))
    return sorted(plages, key=lambda p: p[0])


def _jour_a_dejeuner(jour: str, d) -> bool:
    if jour == "mercredi":
        return False
    if any(t.jour == jour and t.type == "dejeuner" for t in d.horaires):
        return True
    return bool(_plages_dejeuner(d))


def _cell(abrev: str, prof_nom: str, couleurs: dict, salle_nom: str = "") -> str:
    bg, fg = _couleur(abrev, couleurs)
    salle_html = (
        f'<span class="salle">{_libelle_salle_affichage(salle_nom)}</span>'
        if salle_nom else ""
    )
    return (
        f'<div class="cours" style="background:{bg};color:{fg}">'
        f'<div class="cours-top">'
        f'<span class="abrev">{abrev}</span>{salle_html}'
        f'</div>'
        f'<span class="prof">{prof_nom}</span>'
        f'</div>'
    )


def _badge_cours(abrev: str, c_nom: str, couleurs: dict) -> str:
    bg, fg = _couleur(abrev, couleurs)
    return (f'<span class="badge" style="background:{bg};color:{fg}">'
            f'{abrev}/{c_nom}</span>')


def _classe_jour_profs(jour_idx: int) -> str:
    return f"prof-j{jour_idx}"


def _cellule_periode(slots: list[tuple[str, str, str, str]], jour_cls: str,
                     couleurs: dict, periode: str = "matin") -> str:
    cls = f"prof-periode {jour_cls}" + (" aprem-p" if periode == "aprem" else "")
    if not slots:
        return f'<td class="{cls} vide"></td>'
    lignes = "".join(
        f'<div class="prof-slot">'
        f'<span class="prof-heure">{debut}</span>'
        f'<div class="prof-detail">'
        f'{_badge_cours(abrev, c_nom, couleurs)}'
        f'{"<span class=\"prof-salle\">" + _libelle_salle_affichage(s_nom) + "</span>" if s_nom else ""}'
        f'</div>'
        f'</div>'
        for debut, abrev, c_nom, s_nom in slots
    )
    return f'<td class="{cls}">{lignes}</td>'


def _csv_classe(c_id: str, c_nom: str, solution: dict, d, t_map: dict) -> str:
    cours = solution.get(c_id, {})
    debuts_cours = sorted({t.debut for t in d.creneaux}, key=lambda h: h)
    out = io.StringIO()
    w = _csv.writer(out)
    w.writerow(["Heure"] + [JOURS_LABELS[j] for j in JOURS])
    for debut in debuts_cours:
        row = [debut]
        for jour in JOURS:
            t_id = next((t.id for t in d.creneaux if t.jour == jour and t.debut == debut), None)
            if t_id and t_id in cours:
                m_id, p_id, s_id = _unpack_cours(cours[t_id])
                abrev = d.matieres[m_id].abrev if m_id in d.matieres else m_id
                prof = d.profs.get(p_id)
                p_nom = _html_nom_prof(prof) if prof else p_id
                s_nom = _nom_salle(d, c_id, m_id, s_id)
                row.append(f"{abrev} - {p_nom}" + (f" ({s_nom})" if s_nom else ""))
            else:
                row.append("")
        w.writerow(row)
    return out.getvalue()


def pref_satisfaction_dict(solution: dict, d, t_map: dict | None = None) -> dict:
    """Satisfaction des préférences — dict JSON-sérialisable."""
    if t_map is None:
        t_map = {cr.id: cr for cr in d.creneaux}
    planning_jour: dict[str, set] = defaultdict(set)
    planning_debut: dict[str, set] = defaultdict(set)
    for c_id, cours in solution.items():
        for t_id, entry in cours.items():
            m_id, p_id, _ = _unpack_cours(entry)
            t = t_map[t_id]
            planning_jour[p_id].add(t.jour)
            planning_debut[p_id].add(t.debut)
    lignes = []
    ok = ko = 0
    for pref in d.preferences:
        if pref.type not in ("jour_libre", "debut_pas_avant"):
            continue
        prof = d.profs.get(pref.prof_id)
        if not prof:
            continue
        if pref.type == "jour_libre":
            violation = pref.valeur in planning_jour.get(pref.prof_id, set())
        else:
            violation = any(
                debut < pref.valeur for debut in planning_debut.get(pref.prof_id, set())
            )
        if violation:
            ko += 1
        else:
            ok += 1
        lignes.append({
            "satisfait": not violation,
            "prof": prof.nom_complet,
            "type": pref.type,
            "valeur": pref.valeur,
            "priorite": pref.priorite,
        })
    return {"ok": ok, "ko": ko, "total": ok + ko, "lignes": lignes}


def _xlsx_classe_bytes(c_id: str, c_nom: str, solution: dict, d, t_map: dict,
                       couleurs: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return b""
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = c_nom[:31]
    ws.column_dimensions["A"].width = 7
    for letter in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[letter].width = 18
    ws.append(["Heure"] + [JOURS_LABELS[j] for j in JOURS])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A3A5C")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    cours = solution.get(c_id, {})
    for debut in sorted({t.debut for t in d.creneaux}):
        row_data, row_styles = [debut], [None]
        for jour in JOURS:
            t_id = next((t.id for t in d.creneaux if t.jour == jour and t.debut == debut), None)
            if t_id and t_id in cours:
                m_id, p_id, s_id = _unpack_cours(cours[t_id])
                abrev = d.matieres[m_id].abrev if m_id in d.matieres else m_id
                prof  = d.profs.get(p_id)
                s_nom = _nom_salle(d, c_id, m_id, s_id)
                txt = f"{abrev}\n{(prof.nom_complet if prof else p_id)}"
                if s_nom:
                    txt += f"\n{_libelle_salle_affichage(s_nom)}"
                row_data.append(txt)
                bg, _ = _couleur(abrev, couleurs)
                row_styles.append((_excel_rgb(bg), _excel_rgb(_auto_fg(bg))))
            else:
                row_data.append("")
                row_styles.append(None)
        ws.append(row_data)
        ri = ws.max_row
        has_cours = any(row_styles[i] for i in range(1, len(row_styles)))
        ws.row_dimensions[ri].height = _XLSX_ROW_HEIGHT_CLASSE if has_cours else 18
        ws.cell(ri, 1).alignment = Alignment(horizontal="center", vertical="center")
        for col, style in enumerate(row_styles[1:], 2):
            cell = ws.cell(ri, col)
            if style:
                cell.fill = PatternFill("solid", fgColor=style[0])
                cell.font = _xlsx_font(style[1])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb.save(buf)
    return buf.getvalue()


def _xlsx_profs_bytes(solution: dict, d, t_map: dict, couleurs: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return b""
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Professeurs"
    header = ["Professeur"]
    for jour in JOURS:
        header.append(f"{JOURS_LABELS[jour]} matin")
        if jour != "mercredi":
            header.append(f"{JOURS_LABELS[jour]} aprem")
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2D5986")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, len(header) + 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 16
    planning: dict[str, dict] = defaultdict(dict)
    for c_id, cours in solution.items():
        for t_id, entry in cours.items():
            m_id, p_id, s_id = _unpack_cours(entry)
            planning[p_id][t_id] = (c_id, m_id, s_id)
    for p_id in sorted(planning):
        prof    = d.profs[p_id]
        cours_p = planning[p_id]
        par_periode: dict[tuple, list] = defaultdict(list)
        for t_id, (c_id, m_id, s_id) in cours_p.items():
            t     = t_map[t_id]
            abrev = d.matieres[m_id].abrev if m_id in d.matieres else m_id
            cl    = next((c for c in d.classes if c.id == c_id), None)
            s_nom = _nom_salle(d, c_id, m_id, s_id)
            par_periode[(t.jour, "matin" if t.debut < "12:00" else "aprem")].append(
                (t.debut, abrev, cl.nom if cl else c_id, s_nom)
            )
        row_data   = [f"{prof.nom_complet}\n{len(cours_p)}h/{prof.h_contrat}h"]
        row_colors = [None]
        for jour in JOURS:
            for periode in ("matin", "aprem"):
                if jour == "mercredi" and periode == "aprem":
                    continue
                slots = sorted(par_periode.get((jour, periode), []))
                row_data.append("\n".join(
                    f"{a} {c}" + (f"\n{_libelle_salle_affichage(s)}" if s else "")
                    for _, a, c, s in slots
                ) if slots else "")
                if slots:
                    bg, _ = _couleur(slots[0][1], couleurs)
                    row_colors.append(_excel_rgb(bg))
                else:
                    row_colors.append(None)
        ws.append(row_data)
        ri = ws.max_row
        ws.row_dimensions[ri].height = _XLSX_ROW_HEIGHT_PROF
        ws.cell(ri, 1).font = _xlsx_font("000000", size=_XLSX_FONT_SIZE_PROF)
        ws.cell(ri, 1).alignment = Alignment(wrap_text=True, vertical="center")
        for col, hex_bg in enumerate(row_colors[1:], 2):
            cell = ws.cell(ri, col)
            if hex_bg:
                cell.fill = PatternFill("solid", fgColor=hex_bg)
                cell.font = _xlsx_font(_excel_rgb(_auto_fg("#" + hex_bg)), size=_XLSX_FONT_SIZE_PROF)
            else:
                cell.font = _xlsx_font("000000", size=_XLSX_FONT_SIZE_PROF)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb.save(buf)
    return buf.getvalue()


def _grille_classe(c_id: str, c_nom: str, solution: dict, d, t_map: dict,
                   couleurs: dict) -> str:
    plages_dej = _plages_dejeuner(d)
    debuts_cours = sorted({t.debut for t in d.creneaux}, key=lambda h: h)
    debuts_aprem = {t.debut for t in d.creneaux if t.debut >= "13:00"}
    timeline: list[tuple[str, str]] = []
    dej_inseres = False
    for debut in debuts_cours:
        if not dej_inseres and debut in debuts_aprem and plages_dej:
            for p in plages_dej:
                timeline.append((p[0], "dejeuner"))
            dej_inseres = True
        timeline.append((debut, "cours"))
    if not dej_inseres and plages_dej:
        for p in plages_dej:
            timeline.append((p[0], "dejeuner"))

    cours = solution.get(c_id, {})
    dej_map = {p[0]: p for p in plages_dej}

    rows = []
    for debut, kind in timeline:
        if kind == "dejeuner":
            debut_d, fin_d = dej_map[debut]
            cells = [
                f'<td class="heure dejeuner">{_fmt_horaire(debut_d)}<br>'
                f'<small>{_fmt_horaire(fin_d)}</small></td>'
            ]
            for jour in JOURS:
                if _jour_a_dejeuner(jour, d):
                    cells.append(
                        f'<td class="dejeuner">Déjeuner<br>'
                        f'<small>{_fmt_horaire(debut_d)}–{_fmt_horaire(fin_d)}</small></td>'
                    )
                else:
                    cells.append('<td class="absent"></td>')
        else:
            cells = [f'<td class="heure">{debut}</td>']
            for jour in JOURS:
                t_id = next(
                    (t.id for t in d.creneaux if t.jour == jour and t.debut == debut),
                    None,
                )
                if t_id and t_id in cours:
                    m_id, p_id, s_id = _unpack_cours(cours[t_id])
                    abrev = d.matieres[m_id].abrev if m_id in d.matieres else m_id
                    prof = d.profs.get(p_id)
                    p_nom = _html_nom_prof(prof) if prof else p_id
                    s_nom = _nom_salle(d, c_id, m_id, s_id)
                    cells.append(f'<td>{_cell(abrev, p_nom, couleurs, s_nom)}</td>')
                elif t_id:
                    cells.append('<td class="vide"></td>')
                else:
                    cells.append('<td class="absent"></td>')
        rows.append("<tr>" + "".join(cells) + "</tr>")

    header_cells = ['<th>Heure</th>'] + [f'<th>{JOURS_LABELS[j]}</th>' for j in JOURS]
    header = "<tr>" + "".join(header_cells) + "</tr>"

    csv_data  = _csv_classe(c_id, c_nom, solution, d, t_map)
    csv_url   = "data:text/csv;charset=utf-8," + urllib.parse.quote(csv_data)
    _xb       = _xlsx_classe_bytes(c_id, c_nom, solution, d, t_map, couleurs)
    xlsx_url  = ("data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
                 + _base64.b64encode(_xb).decode()) if _xb else ""
    xlsx_btn  = f'<a class="btn-action" href="{xlsx_url}" download="{c_nom}.xlsx">📊 XLSX</a>' if xlsx_url else ""

    nb_cours = len(cours)
    return f"""
<div class="tab-actions">
  <a class="btn-action" href="{csv_url}" download="{c_nom}.csv">📥 CSV</a>
  {xlsx_btn}
  <button class="btn-action" onclick="printCurrentTab()" data-tooltip="Imprimer l'onglet classe actif">🖨 Imprimer</button>
</div>
<table class="grille">
  <thead>{header}</thead>
  <tbody>{"".join(rows)}</tbody>
</table>
<div class="stat">{nb_cours} cours / semaine</div>"""


def _tableau_profs(solution: dict, d, t_map: dict, couleurs: dict) -> str:
    planning: dict[str, dict] = defaultdict(dict)
    for c_id, cours in solution.items():
        for t_id, entry in cours.items():
            m_id, p_id, s_id = _unpack_cours(entry)
            planning[p_id][t_id] = (c_id, m_id, s_id)

    hdr_jours = ['<th rowspan="2" class="prof-nom-h">Professeur</th>']
    for jour_idx, jour in enumerate(JOURS):
        jour_cls = _classe_jour_profs(jour_idx)
        colspan = 1 if jour == "mercredi" else 2
        hdr_jours.append(
            f'<th colspan="{colspan}" class="prof-jour-h {jour_cls}">'
            f'{JOURS_LABELS[jour]}</th>'
        )
    hdr_periodes = []
    for jour_idx, jour in enumerate(JOURS):
        jour_cls = _classe_jour_profs(jour_idx)
        hdr_periodes.append(f'<th class="prof-per-h {jour_cls}">matin</th>')
        if jour != "mercredi":
            hdr_periodes.append(f'<th class="prof-per-h {jour_cls}">aprem</th>')

    hdr_repeat = (
        f'<tr class="hdr-repeat screen-only">{"".join(hdr_jours)}</tr>'
        f'<tr class="hdr-repeat screen-only">{"".join(hdr_periodes)}</tr>'
    )

    lignes = []
    for i, p_id in enumerate(sorted(planning)):
        if i > 0 and i % 10 == 0:
            lignes.append(hdr_repeat)

        prof = d.profs[p_id]
        cours = planning[p_id]
        total = len(cours)
        contrat = prof.h_contrat

        par_periode: dict[tuple[str, str], list[tuple[str, str, str, str]]] = defaultdict(list)
        for t_id, (c_id, m_id, s_id) in cours.items():
            t = t_map[t_id]
            abrev = d.matieres[m_id].abrev if m_id in d.matieres else m_id
            classe = next((c for c in d.classes if c.id == c_id), None)
            c_nom = classe.nom if classe else c_id
            s_nom = _nom_salle(d, c_id, m_id, s_id)
            if t.debut in MATIN_DEBUTS:
                periode = "matin"
            elif t.debut in APREM_DEBUTS:
                periode = "aprem"
            else:
                periode = "matin" if t.debut < "12:00" else "aprem"
            par_periode[(t.jour, periode)].append((t.debut, abrev, c_nom, s_nom))

        cells = [
            f'<td class="prof-nom">{_html_nom_prof(prof)}'
            f'<br><small>{total}h assignées / {contrat}h contrat</small></td>'
        ]
        for jour_idx, jour in enumerate(JOURS):
            jour_cls = _classe_jour_profs(jour_idx)
            for periode in ("matin", "aprem"):
                if jour == "mercredi" and periode == "aprem":
                    continue
                slots = sorted(par_periode.get((jour, periode), []), key=lambda s: s[0])
                cells.append(_cellule_periode(slots, jour_cls, couleurs, periode))

        tr_cls = ' class="prof-row-combleur"' if _est_combleur_nom(prof.nom_complet) else ""
        lignes.append(f"<tr{tr_cls}>" + "".join(cells) + "</tr>")

    return f"""
<table class="profs">
  <thead>
    <tr>{"".join(hdr_jours)}</tr>
    <tr>{"".join(hdr_periodes)}</tr>
  </thead>
  <tbody>{"".join(lignes)}</tbody>
</table>"""


_PRIO_RANK = {"contrainte": 0, "haute": 1, "moyenne": 2, "basse": 3}


def _rapport_preferences(solution: dict, d, t_map: dict) -> str:
    planning_jour: dict[str, set[str]] = defaultdict(set)
    planning_debut: dict[str, set[str]] = defaultdict(set)
    for c_id, cours in solution.items():
        for t_id, entry in cours.items():
            m_id, p_id, _ = _unpack_cours(entry)
            t = t_map[t_id]
            planning_jour[p_id].add(t.jour)
            planning_debut[p_id].add(t.debut)

    lignes = []
    ok = ko = 0
    for pref in d.preferences:
        if pref.type not in ("jour_libre", "debut_pas_avant"):
            continue
        prof = d.profs.get(pref.prof_id)
        if not prof:
            continue

        if pref.type == "jour_libre":
            violation = pref.valeur in planning_jour.get(pref.prof_id, set())
        else:
            violation = any(
                debut < pref.valeur
                for debut in planning_debut.get(pref.prof_id, set())
            )

        if violation:
            statut = '❌ KO'
            cls = "ko"
            ko += 1
        else:
            statut = '✅ OK'
            cls = "ok"
            ok += 1

        prio_cls = "prio-haute" if pref.priorite == "haute" else "prio-moyenne"
        stat_num = "0" if violation else "1"
        prio_num = str(_PRIO_RANK.get(pref.priorite, 9))
        lignes.append(
            f'<tr class="{cls}"'
            f' data-statut="{stat_num}"'
            f' data-prof="{_html.escape(prof.nom_complet, quote=True)}"'
            f' data-type="{_html.escape(pref.type, quote=True)}"'
            f' data-valeur="{_html.escape(pref.valeur, quote=True)}"'
            f' data-prio="{prio_num}">'
            f'<td>{statut}</td>'
            f'<td>{_html_nom_prof(prof)}</td>'
            f'<td>{pref.type}</td>'
            f'<td>{pref.valeur}</td>'
            f'<td class="{prio_cls}">{pref.priorite}</td>'
            f'</tr>'
        )

    total = ok + ko
    bilan = (f'<p class="bilan ok">✅ {ok}/{total} préférences satisfaites</p>'
             if ko == 0
             else f'<p class="bilan ko">⚠ {ok}/{total} satisfaites — {ko} non validée(s)</p>')

    header = (
        "<tr>"
        '<th class="prefs-sort" data-col="statut">Statut ↕</th>'
        '<th class="prefs-sort" data-col="prof">Professeur ↕</th>'
        '<th class="prefs-sort" data-col="type">Type ↕</th>'
        '<th class="prefs-sort" data-col="valeur">Valeur ↕</th>'
        '<th class="prefs-sort" data-col="prio">Priorité ↕</th>'
        "</tr>"
    )
    return f"""
{bilan}
<table class="prefs">
  <thead>{header}</thead>
  <tbody>{"".join(lignes)}</tbody>
</table>"""


def generer_html(solution: dict, d, output: Path, score: float = 0.0,
                 temps: float = 0.0, propagations: int = 0) -> None:
    t_map = {cr.id: cr for cr in d.creneaux}
    couleurs = _build_couleurs(d)

    tabs_btn = []
    tabs_content = []
    for i, classe in enumerate(d.classes):
        active = "active" if i == 0 else ""
        spec = f" <em>({classe.specificite})</em>" if classe.specificite else ""
        tabs_btn.append(
            f'<button class="tab-btn {active}" onclick="showTab(\'{classe.id}\')">'
            f'{classe.nom}{spec}</button>'
        )
        display    = "block" if i == 0 else "none"
        active_cls = " active" if i == 0 else ""
        tabs_content.append(
            f'<div id="tab-{classe.id}" class="tab-pane{active_cls}" style="display:{display}">'
            f'<h2 class="classe-titre">{classe.nom} — {classe.nb_eleves} élèves</h2>'
            + _grille_classe(classe.id, classe.nom, solution, d, t_map, couleurs)
            + '</div>'
        )

    _xp            = _xlsx_profs_bytes(solution, d, t_map, couleurs)
    xlsx_profs_url = ("data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
                      + _base64.b64encode(_xp).decode()) if _xp else ""
    xlsx_profs_btn = (f'<a class="btn-action" href="{xlsx_profs_url}" download="planning_profs.xlsx">📊 XLSX</a>'
                      if xlsx_profs_url else "")

    legende = " ".join(
        f'<span class="leg" style="background:{bg};color:{fg}">{abrev}</span>'
        for abrev, (bg, fg) in couleurs.items()
        if abrev in {m.abrev for m in d.matieres.values()}
    )

    html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Emploi du temps</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', sans-serif; background: #f4f5f7; color: #333; }}

  header {{ background: #1a3a5c; color: #fff; padding: 14px 24px;
            display: flex; align-items: center; gap: 16px; }}
  header h1 {{ font-size: 1.2em; flex: 1; }}
  header p  {{ font-size: .82em; opacity: .8; }}
  .btn-print {{
    background: rgba(255,255,255,.15); border: 1px solid rgba(255,255,255,.4);
    color: #fff; padding: 6px 14px; border-radius: 4px; cursor: pointer;
    font-size: .85em; white-space: nowrap;
  }}
  .btn-print:hover {{ background: rgba(255,255,255,.28); }}

  .container {{ max-width: 1600px; margin: 0 auto; padding: 20px; }}

  .tabs {{ display: flex; flex-wrap: wrap; gap: 4px; margin-bottom: 16px; }}
  .tab-btn {{
    padding: 6px 12px; border: none; border-radius: 4px;
    background: #dde3ea; cursor: pointer; font-size: .82em; transition: background .15s;
  }}
  .tab-btn.active, .tab-btn:hover {{ background: #1a3a5c; color: #fff; }}
  .tab-btn em {{ font-style: normal; font-size: .85em; opacity: .75; }}

  .tab-actions {{ display: flex; gap: 8px; margin-bottom: 10px; }}
  .btn-action {{
    display: inline-block; padding: 5px 12px; border-radius: 4px;
    background: #e8f0fa; color: #1a3a5c; text-decoration: none;
    font-size: .82em; border: 1px solid #c5d5e8; cursor: pointer;
  }}
  .btn-action:hover {{ background: #d0e3f7; }}

  h2 {{ font-size: 1.1em; margin-bottom: 10px; color: #1a3a5c; }}
  .grille {{ border-collapse: collapse; width: 100%; font-size: .82em; }}
  .grille th {{ background: #1a3a5c; color: #fff; padding: 6px 8px; text-align: center; }}
  .grille td {{ border: 1px solid #ddd; padding: 4px; vertical-align: top;
               min-width: 110px; height: 46px; }}
  .grille .heure {{ text-align: right; font-size: .78em; color: #888;
                    background: #f9f9f9; padding-right: 6px; min-width: 54px; }}
  .grille .vide {{ background: #fafafa; }}
  .grille .absent {{ background: #f0f0f0; opacity: .4; }}
  .grille .dejeuner {{ background: #fff8e6; color: #7a5c00; text-align: center;
                      font-size: .82em; font-weight: 600; vertical-align: middle; }}
  .grille .heure.dejeuner {{ background: #fff3cc; color: #7a5c00; }}

  .cours {{ border-radius: 4px; padding: 3px 5px; font-size: .88em;
            display: flex; flex-direction: column;
            justify-content: center; height: 100%; line-height: 1.15; }}
  .cours-top {{ display: flex; justify-content: space-between; align-items: baseline;
                width: 100%; min-width: 0; }}
  .cours .abrev {{ font-weight: 700; flex-shrink: 0; }}
  .cours .salle {{ font-size: .88em; font-weight: 400; font-style: normal;
                  text-align: right; margin-left: 6px; opacity: .92;
                  overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
  .cours .prof  {{ font-size: .7em; opacity: .92; margin-top: 1px; }}
  .prof-detail {{ display: flex; flex-direction: row; align-items: flex-start;
                  justify-content: space-between; gap: 4px; min-width: 0; flex: 1; }}
  .prof-detail .badge {{ flex-shrink: 0; }}
  .prof-salle {{ font-size: .78em; color: #444; font-weight: 500; line-height: 1.2;
                 text-align: right; flex: 1; min-width: 0;
                 white-space: normal; word-break: break-word; }}

  .stat {{ font-size: .8em; color: #666; margin-top: 6px; text-align: right; }}

  .legende {{ display: flex; flex-wrap: wrap; gap: 6px; margin: 14px 0; }}
  .leg {{ padding: 3px 8px; border-radius: 3px; font-size: .78em; font-weight: 600; }}

  .profs {{ border-collapse: collapse; width: 100%; font-size: .76em; margin-top: 10px; }}
  .profs th {{ background: #2d5986; color: #fff; padding: 5px 6px; text-align: center; }}
  .profs td {{ border: 1px solid #ddd; padding: 3px 4px; vertical-align: top; }}
  .profs thead {{ display: table-header-group; }}
  /* Répétition écran (tous les 10 profs) — mêmes couleurs que thead */
  .profs tr.hdr-repeat th.prof-nom-h {{ background: #2d5986; }}
  .profs tr.hdr-repeat th.prof-j0 {{ background: #2d5986; color: #fff; }}
  .profs tr.hdr-repeat th.prof-j1 {{ background: #2a7a5f; color: #fff; }}
  .profs tr.hdr-repeat th.prof-j2 {{ background: #553a8a; color: #fff; }}
  .profs tr.hdr-repeat th.prof-j3 {{ background: #8a5a25; color: #fff; }}
  .profs tr.hdr-repeat th.prof-j4 {{ background: #3a7035; color: #fff; }}
  .prof-nom-h {{ min-width: 140px; text-align: left !important; }}
  .prof-jour-h {{ font-size: .9em; }}
  .prof-per-h {{ font-size: .75em; font-weight: 500; opacity: .9; }}
  .prof-nom {{ background: #eef2f7; font-weight: 600; min-width: 140px; }}
  .prof-combleur {{ color: #c0392b; font-weight: 700; cursor: help;
    border-bottom: 1px dotted #c0392b; }}
  tr.prof-row-combleur td.prof-nom {{ background: #fdecea; }}
  .prof-periode {{ min-width: 108px; }}
  /* 5 couleurs par jour — matin (clair) / aprem (légèrement plus foncé) */
  .profs td.prof-j0        {{ background: #dce8f8; }}
  .profs td.prof-j0.aprem-p{{ background: #c2d4ee; }}
  .profs td.prof-j1        {{ background: #d8f0e8; }}
  .profs td.prof-j1.aprem-p{{ background: #bde2d5; }}
  .profs td.prof-j2        {{ background: #ece8f8; }}
  .profs td.prof-j2.aprem-p{{ background: #d9d0f0; }}
  .profs td.prof-j3        {{ background: #f8ede0; }}
  .profs td.prof-j3.aprem-p{{ background: #eedcc8; }}
  .profs td.prof-j4        {{ background: #dff2dc; }}
  .profs td.prof-j4.aprem-p{{ background: #c8e5c3; }}
  /* En-têtes jour */
  .profs th.prof-j0 {{ background: #2d5986; color:#fff; }}
  .profs th.prof-j1 {{ background: #2a7a5f; color:#fff; }}
  .profs th.prof-j2 {{ background: #553a8a; color:#fff; }}
  .profs th.prof-j3 {{ background: #8a5a25; color:#fff; }}
  .profs th.prof-j4 {{ background: #3a7035; color:#fff; }}
  .profs td.prof-periode.vide {{ filter: brightness(1.03); }}
  .profs td.prof-periode.absent {{ opacity: .45; filter: brightness(.97); }}
  .prof-slot {{ display: flex; align-items: flex-start; gap: 4px; margin: 2px 0; line-height: 1.25; }}
  .prof-heure {{ font-size: .78em; color: #666; font-weight: 700;
                min-width: 38px; flex-shrink: 0; }}
  .badge {{ display: inline-block; border-radius: 3px; padding: 2px 5px;
            font-size: .85em; white-space: nowrap; }}

  .prefs {{ border-collapse: collapse; width: 100%; font-size: .82em; margin-top: 10px; }}
  .prefs th {{ background: #2d5986; color: #fff; padding: 6px 8px; }}
  .prefs th.prefs-sort {{ cursor: pointer; user-select: none; }}
  .prefs th.prefs-sort:hover {{ background: #234a6e; }}
  .prefs td {{ border: 1px solid #ddd; padding: 5px 8px; }}
  .prefs .ok  {{ background: #f0fff4; }}
  .prefs .ko  {{ background: #fff5f5; }}
  .prio-haute   {{ color: #c0392b; font-weight: 700; }}
  .prio-moyenne {{ color: #e67e22; }}
  .bilan {{ margin-bottom: 10px; font-weight: 600; padding: 8px 12px;
            border-radius: 4px; }}
  .bilan.ok {{ background: #d4edda; color: #155724; }}
  .bilan.ko {{ background: #f8d7da; color: #721c24; }}

  .section {{ background: #fff; border-radius: 8px; padding: 20px;
              box-shadow: 0 1px 3px rgba(0,0,0,.1); margin-bottom: 20px; }}
  .section-title {{ font-size: 1.05em; font-weight: 700; color: #1a3a5c;
                    border-bottom: 2px solid #1a3a5c; padding-bottom: 6px;
                    margin-bottom: 14px; }}
  .meta {{ font-size: .8em; color: #888; margin-top: 6px; }}

  .section-actions {{ display: flex; gap: 8px; margin-bottom: 12px; }}

  @media print {{
    * {{
      -webkit-print-color-adjust: exact !important;
      print-color-adjust: exact !important;
    }}
    .tab-actions, .section-actions {{ display: none !important; }}
    .section {{ box-shadow: none; }}
    body {{ background: #fff; }}
    .tabs {{ display: none !important; }}
    .tab-pane {{ display: block !important; page-break-after: always; }}

    body.printing-class header {{ display: none !important; }}
    body.printing-class .section-title,
    body.printing-class .legende,
    body.printing-class .stat {{ display: none !important; }}
    body.printing-class .container {{ padding: 0; max-width: none; }}
    body.printing-class .classes-section {{
      padding: 0; margin: 0; box-shadow: none; border-radius: 0;
    }}
    body.printing-class .classe-titre {{
      font-size: 14pt; font-weight: 700; color: #000;
      margin: 0 0 10px 0; page-break-after: avoid;
    }}
    body.printing-class .grille th {{
      background: #1a3a5c !important; color: #fff !important;
    }}
    body.printing-class .grille .dejeuner,
    body.printing-class .grille .heure.dejeuner {{
      background: #fff8e6 !important; color: #7a5c00 !important;
    }}

    body.printing-class .tab-pane {{ display: none !important; }}
    body.printing-class .tab-pane.active {{ display: block !important; page-break-after: auto; }}
    body.printing-class .section:not(.classes-section),
    body.printing-class .meta {{ display: none !important; }}

    body.printing-profs .section:not(.profs-section),
    body.printing-profs .meta {{ display: none !important; }}
    body.printing-profs header,
    body.printing-profs .section-title,
    body.printing-profs .section-actions {{ display: none !important; }}
    body.printing-profs .container {{ padding: 0; max-width: none; }}
    body.printing-profs .profs-section {{
      padding: 0; margin: 0; box-shadow: none; border-radius: 0;
    }}
    body.printing-profs .profs thead {{ display: table-header-group; }}
    body.printing-profs .profs tbody tr.hdr-repeat,
    body.printing-profs .profs tbody tr.screen-only {{ display: none !important; }}
    body.printing-profs .profs th.prof-nom-h {{ background: #2d5986 !important; color: #fff !important; }}
    body.printing-profs .profs th.prof-j0 {{ background: #2d5986 !important; color: #fff !important; }}
    body.printing-profs .profs th.prof-j1 {{ background: #2a7a5f !important; color: #fff !important; }}
    body.printing-profs .profs th.prof-j2 {{ background: #553a8a !important; color: #fff !important; }}
    body.printing-profs .profs th.prof-j3 {{ background: #8a5a25 !important; color: #fff !important; }}
    body.printing-profs .profs th.prof-j4 {{ background: #3a7035 !important; color: #fff !important; }}
    body.printing-profs .profs td.prof-j0        {{ background: #dce8f8 !important; }}
    body.printing-profs .profs td.prof-j0.aprem-p{{ background: #c2d4ee !important; }}
    body.printing-profs .profs td.prof-j1        {{ background: #d8f0e8 !important; }}
    body.printing-profs .profs td.prof-j1.aprem-p{{ background: #bde2d5 !important; }}
    body.printing-profs .profs td.prof-j2        {{ background: #ece8f8 !important; }}
    body.printing-profs .profs td.prof-j2.aprem-p{{ background: #d9d0f0 !important; }}
    body.printing-profs .profs td.prof-j3        {{ background: #f8ede0 !important; }}
    body.printing-profs .profs td.prof-j3.aprem-p{{ background: #eedcc8 !important; }}
    body.printing-profs .profs td.prof-j4        {{ background: #dff2dc !important; }}
    body.printing-profs .profs td.prof-j4.aprem-p{{ background: #c8e5c3 !important; }}
    body.printing-profs .prof-nom {{ background: #eef2f7 !important; }}
    body.printing-profs .profs tr {{ page-break-inside: avoid; }}
  }}
</style>
</head>
<body>

<header>
  <div style="flex:1">
    <h1>📅 Emploi du temps</h1>
    <p>Résolution en {temps:.2f}s &nbsp;|&nbsp; Imperfections : {score:.0f}{f" &nbsp;|&nbsp; {propagations:,} propagations OR-Tools".replace(",", " ") if propagations else ""}</p>
  </div>
</header>

<div class="container">

  <div class="section classes-section">
    <div class="section-title">📚 Emploi du temps par classe</div>
    <div class="legende">{legende}</div>
    <div class="tabs">{"".join(tabs_btn)}</div>
    {"".join(tabs_content)}
  </div>

  <div class="section profs-section">
    <div class="section-title">👩‍🏫 Planning des professeurs</div>
    <div class="section-actions">
      {xlsx_profs_btn}
      <button class="btn-action" onclick="printProfs()" data-tooltip="Imprimer le planning professeurs">🖨 Imprimer</button>
    </div>
    {_tableau_profs(solution, d, t_map, couleurs)}
  </div>

  <div class="section">
    <div class="section-title">✅ Satisfaction des préférences</div>
    {_rapport_preferences(solution, d, t_map)}
  </div>

  <p class="meta">
    Généré par <strong>TimeTabler</strong> (OR-Tools CP-SAT) ·
    {sum(c.nb_eleves for c in d.classes)} élèves · {len(d.classes)} classes ·
    {len(d.profs)} professeurs · {len(d.creneaux)} créneaux cours / semaine
  </p>
</div>

<script>
function showTab(id) {{
  document.querySelectorAll('.tab-pane').forEach(p => {{ p.style.display = 'none'; p.classList.remove('active'); }});
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  const pane = document.getElementById('tab-' + id);
  pane.style.display = 'block';
  pane.classList.add('active');
  event.target.classList.add('active');
}}
function printCurrentTab() {{
  document.body.classList.add('printing-class');
  window.print();
  document.body.classList.remove('printing-class');
}}
function printProfs() {{
  document.body.classList.add('printing-profs');
  var detached = [];
  document.querySelectorAll('.profs tbody tr.hdr-repeat').forEach(function(tr) {{
    detached.push({{ tr: tr, parent: tr.parentNode, next: tr.nextSibling }});
    tr.parentNode.removeChild(tr);
  }});
  window.print();
  detached.forEach(function(d) {{
    if (d.next) d.parent.insertBefore(d.tr, d.next);
    else d.parent.appendChild(d.tr);
  }});
  document.body.classList.remove('printing-profs');
}}
function initPrefsSort() {{
  var table = document.querySelector('table.prefs');
  if (!table) return;
  var st = {{ col: 'statut', dir: 1 }};
  var labels = {{ statut: 'Statut', prof: 'Professeur', type: 'Type', valeur: 'Valeur', prio: 'Priorité' }};
  function cmp(a, b) {{
    var va = a.dataset[st.col] || '';
    var vb = b.dataset[st.col] || '';
    var c = 0;
    if (st.col === 'statut' || st.col === 'prio') c = (parseInt(va, 10) || 0) - (parseInt(vb, 10) || 0);
    else c = String(va).localeCompare(String(vb), 'fr');
    return c * st.dir;
  }}
  function sortRows() {{
    var tbody = table.querySelector('tbody');
    Array.from(tbody.querySelectorAll('tr')).sort(cmp).forEach(function(r) {{ tbody.appendChild(r); }});
    table.querySelectorAll('th.prefs-sort').forEach(function(th) {{
      var col = th.dataset.col;
      var arrow = col === st.col ? (st.dir > 0 ? ' ▲' : ' ▼') : ' ↕';
      th.textContent = labels[col] + arrow;
    }});
  }}
  table.querySelectorAll('th.prefs-sort').forEach(function(th) {{
    th.onclick = function() {{
      var col = th.dataset.col;
      if (st.col === col) st.dir *= -1;
      else {{ st.col = col; st.dir = 1; }}
      sortRows();
    }};
  }});
  sortRows();
}}
initPrefsSort();
</script>
<script src="https://dnavatar.org/_interfaces/kill-native-tooltip.js"></script>
<script src="https://dnavatar.org/_interfaces/tooltips.js"></script>

</body>
</html>"""

    output.write_text(html, encoding="utf-8")
    print(f"  HTML généré : {output}")
